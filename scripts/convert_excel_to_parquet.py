"""
convert_excel_to_parquet.py

통합 문서1.xlsx -> long-format Parquet 변환
- market_data_long.parquet  (SSC: 수정주가, 시가총액, 상장주식수)
- financial_data_long.parquet (NFS-IFRS: 자산총계, 자본총계, 자본잉여금, 이익잉여금, 당기순이익, 기타포괄이익)

Excel 구조:
  Row 26: 헤더 (col14=코드, col15=종목명, col16=구분, col17=아이템코드, col18=아이템명, col19=기간구분, col20~=날짜)
  Row 27~: 데이터
"""

import os
import sys
import pandas as pd
import numpy as np
import openpyxl
from collections import defaultdict

# ============================================================
# 경로 설정
# ============================================================
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(os.path.dirname(BASE_DIR), '통합 문서1.xlsx')
DATA_DIR = os.path.join(BASE_DIR, 'data')
os.makedirs(DATA_DIR, exist_ok=True)

# ============================================================
# Excel 구조 파라미터
# ============================================================
HEADER_ROW = 26
DATA_START_ROW = 27
CODE_COL = 14       # 종목코드 (col O)
NAME_COL = 15       # 종목명 (col P)
DIV_COL = 16        # 구분 (col Q)
ITEMCODE_COL = 17   # 아이템코드 (col R)
ITEMNAME_COL = 18   # 아이템명 (col S)
PERIOD_COL = 19     # 기간구분 (col T)
DATE_COL_START = 20 # 날짜 컬럼 시작 (col U)

# ============================================================
# 필요한 아이템코드
# ============================================================
MARKET_ITEMS = {
    'S410000700',  # 수정주가
    'S410001250',  # 시가총액
    'S420003800',  # 상장주식수
}

FINANCIAL_ITEMS = {
    'M000901001',  # 자산총계
    'M000903001',  # 자본총계 (BE)
    'M000903012',  # 자본잉여금
    'M000903016',  # 이익잉여금
    'M000908001',  # 당기순이익
    'M000908007',  # 기타포괄이익
}

ALL_NEEDED_ITEMS = MARKET_ITEMS | FINANCIAL_ITEMS


def safe_float(v):
    """안전한 float 변환"""
    if v is None:
        return np.nan
    try:
        return float(v)
    except (ValueError, TypeError):
        return np.nan


def parse_date_header(val):
    """날짜 헤더 -> YYYY-MM 문자열"""
    if val is None:
        return None
    if hasattr(val, 'year') and hasattr(val, 'month'):
        return f"{val.year:04d}-{val.month:02d}"
    if isinstance(val, str):
        try:
            # YYYY-MM 형식
            return val.strip()[:7]
        except Exception:
            return None
    return None


def read_excel_to_long():
    """
    openpyxl read_only 모드로 Excel을 읽어 long-format 레코드 생성
    메모리 효율을 위해 iter_rows 사용, 한 행씩 처리
    """
    print(f"[1/3] Excel 열기: {os.path.basename(EXCEL_PATH)}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb.active

    # 날짜 헤더 파싱
    date_headers = []
    for i, row in enumerate(ws.iter_rows(max_row=HEADER_ROW), 1):
        if i == HEADER_ROW:
            vals = [cell.value for cell in row]
            raw_dates = vals[DATE_COL_START:]
            for j, h in enumerate(raw_dates):
                d = parse_date_header(h)
                if d is not None:
                    date_headers.append((j, d))
            break

    print(f"  날짜 컬럼: {len(date_headers)}개 ({date_headers[0][1]} ~ {date_headers[-1][1]})")

    # Long-format 레코드 수집
    market_records = []
    financial_records = []

    row_num = 0
    for row in ws.iter_rows(min_row=DATA_START_ROW):
        row_num += 1
        vals = [cell.value for cell in row]

        if len(vals) <= ITEMCODE_COL:
            continue

        code = vals[CODE_COL]
        name = vals[NAME_COL]
        item_code = vals[ITEMCODE_COL]

        if code is None or item_code is None:
            continue

        code_str = str(code).strip()
        name_str = str(name).strip() if name is not None else ''
        item_code_str = str(item_code).strip()

        if item_code_str not in ALL_NEEDED_ITEMS:
            continue

        data_vals = vals[DATE_COL_START:]
        target_list = market_records if item_code_str in MARKET_ITEMS else financial_records

        for col_offset, date_str in date_headers:
            if col_offset < len(data_vals):
                fv = safe_float(data_vals[col_offset])
                if not np.isnan(fv):
                    target_list.append({
                        'code': code_str,
                        'name': name_str,
                        'item_code': item_code_str,
                        'date': date_str,
                        'value': fv,
                    })

        if row_num % 2000 == 0:
            sys.stdout.write(
                f"\r  행 {row_num} 처리 중 (market: {len(market_records)}, financial: {len(financial_records)})"
            )
            sys.stdout.flush()

    wb.close()
    sys.stdout.write(f"\r  완료 - 행 {row_num}개, market: {len(market_records)}, financial: {len(financial_records)}\n")
    sys.stdout.flush()

    return market_records, financial_records


def save_parquet(records, output_path, dataset_name):
    """레코드 리스트를 Parquet로 저장"""
    if not records:
        print(f"  [SKIP] {dataset_name} - 데이터 없음")
        return

    df = pd.DataFrame(records)
    df['date'] = pd.to_datetime(df['date'] + '-01', format='%Y-%m-%d')
    df = df.sort_values(['code', 'item_code', 'date']).reset_index(drop=True)

    # Parquet 저장 (pyarrow 엔진)
    df.to_parquet(output_path, engine='pyarrow', index=False)

    print(f"  [OK] {dataset_name}")
    print(f"       Rows: {len(df):,}")
    print(f"       Stocks: {df['code'].nunique()}")
    print(f"       Items: {df['item_code'].nunique()}")
    print(f"       Dates: {df['date'].min().strftime('%Y-%m')} ~ {df['date'].max().strftime('%Y-%m')}")
    print(f"       Path: {output_path}")


def main():
    print("=" * 60)
    print("통합 문서1.xlsx -> Long-format Parquet 변환")
    print("=" * 60)

    # 1. Excel 읽기
    market_records, financial_records = read_excel_to_long()

    # 2. 저장
    print("\n[2/3] Parquet 저장...")
    market_path = os.path.join(DATA_DIR, 'market_data_long.parquet')
    financial_path = os.path.join(DATA_DIR, 'financial_data_long.parquet')

    save_parquet(market_records, market_path, 'market_data_long')
    save_parquet(financial_records, financial_path, 'financial_data_long')

    # 3. 검증
    print("\n[3/3] 검증...")
    for path, name, expected_items in [
        (market_path, 'market_data_long', MARKET_ITEMS),
        (financial_path, 'financial_data_long', FINANCIAL_ITEMS),
    ]:
        if not os.path.exists(path):
            print(f"  [FAIL] {name} 파일 없음")
            continue

        df = pd.read_parquet(path)
        actual_items = set(df['item_code'].unique())
        missing = expected_items - actual_items
        extra = actual_items - expected_items

        print(f"  {name}:")
        print(f"    Rows: {len(df):,}")
        print(f"    Stocks: {df['code'].nunique()}")
        print(f"    Items: {actual_items}")
        if missing:
            print(f"    [WARN] Missing items: {missing}")
        if extra:
            print(f"    [WARN] Extra items: {extra}")

        # 삼성전자 데이터 존재 확인
        ss = df[df['code'] == 'A005930']
        print(f"    Samsung (A005930): {ss['item_code'].nunique()} items, {ss['date'].min().strftime('%Y-%m')}~{ss['date'].max().strftime('%Y-%m')}")

    print("\n" + "=" * 60)
    print("변환 완료")
    print("=" * 60)


if __name__ == '__main__':
    main()
